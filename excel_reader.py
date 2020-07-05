import openpyxl as openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.formatting import Rule
from openpyxl.styles import Font, PatternFill, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl import formatting, styles
from openpyxl.styles import NamedStyle


wb2 = load_workbook('tsla.xlsx')
ws2 = wb2.active

wb = Workbook()
ws_data = wb.active

row_num = 0
for row in ws2.iter_rows(max_col=1, values_only=True):
    row_num += 1
    for cell in row:
        value = cell
        ws_data.cell(row=row_num, column=1, value=value)
date_style = NamedStyle(name='datetime', number_format='YYYY-MM-DD')
row_num = 0
for row in ws_data.iter_rows(min_col=1,max_col=1):
    row_num += 1
    ws_data[f'A{row_num}'].style = date_style
ws_data.column_dimensions['A'].width = 11
date_column = 'A'

row_num = 0
for row in ws2.iter_rows(min_col=6, max_col=6, values_only=True):
    row_num += 1
    for cell in row:
        value = cell
        ws_data.cell(row=row_num, column=2, value=value)
adjclose_column = 'B'

row_num = 0
for row in ws2.iter_rows(min_col=7, max_col=7, values_only=True):
    row_num += 1
    for cell in row:
        value = cell
        ws_data.cell(row=row_num, column=3, value=value)
volume_column = 'C'

for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_9 = row_num - 9
        if row_num_minus_9 < 1:
            pass
        elif row_num_minus_9 >= 1:
            moving_average = f'=(SUM(B{row_num}:B{row_num_minus_9})/10)'
            ws_data[f'D{row_num}'] = moving_average
ws_data['D1'] = '10D MA'
ma10d_column = 'D'

current_column = 'E'
for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_9 = row_num - 9
        if row_num_minus_9 < 1:
            pass
        elif row_num_minus_9 >= 1:
            percent_difference = f'=({adjclose_column}{row_num}/{ma10d_column}{row_num})-1'
            ws_data[f'{current_column}{row_num}'] = percent_difference
            ws_data[f'{current_column}{row_num}'].number_format = '0.00%'
ws_data[f'{current_column}1'] = '% diff 10D MA'
rule = ColorScaleRule(start_type='percentile', start_value=2, start_color='ff0000', mid_type='percentile', mid_value=50, mid_color='FFFFFF', end_type='percentile', end_value=98, end_color='FF00AA00')
ws_data.conditional_formatting.add(f'{current_column}1:{current_column}{row_num}', rule)
ma10d_perc_diff_column = 'E'

for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_19 = row_num - 19
        if row_num_minus_19 < 1:
            pass
        elif row_num_minus_19 >= 1:
            moving_average = f'=(SUM(B{row_num}:B{row_num_minus_19})/20)'
            ws_data[f'F{row_num}'] = moving_average
ws_data['F1'] = '20D MA'
ma20d_column = 'F'

current_column = 'G'
for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_19 = row_num - 19
        if row_num_minus_19 < 1:
            pass
        elif row_num_minus_19 >= 1:
            percent_difference = f'=({adjclose_column}{row_num}/{ma20d_column}{row_num})-1'
            ws_data[f'{current_column}{row_num}'] = percent_difference
            ws_data[f'{current_column}{row_num}'].number_format = '0.00%'
ws_data[f'{current_column}1'] = '% diff 20D MA'
rule = ColorScaleRule(start_type='percentile', start_value=2, start_color='ff0000', mid_type='percentile', mid_value=50, mid_color='FFFFFF', end_type='percentile', end_value=98, end_color='FF00AA00')
ws_data.conditional_formatting.add(f'{current_column}1:{current_column}{row_num}', rule)
ma20d_perc_diff_column = 'G'

for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_49 = row_num - 49
        if row_num_minus_49 < 1:
            pass
        elif row_num_minus_49 >= 1:
            moving_average = f'=(SUM(B{row_num}:B{row_num_minus_49})/50)'
            ws_data[f'H{row_num}'] = moving_average
ws_data['H1'] = '50D MA'
ma50d_column = 'H'

current_column = 'I'
for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_49 = row_num - 49
        if row_num_minus_49 < 1:
            pass
        elif row_num_minus_49 >= 1:
            percent_difference = f'=({adjclose_column}{row_num}/{ma50d_column}{row_num})-1'
            ws_data[f'{current_column}{row_num}'] = percent_difference
            ws_data[f'{current_column}{row_num}'].number_format = '0.00%'
ws_data[f'{current_column}1'] = '% diff 50D MA'
rule = ColorScaleRule(start_type='percentile', start_value=2, start_color='ff0000', mid_type='percentile', mid_value=50, mid_color='FFFFFF', end_type='percentile', end_value=98, end_color='FF00AA00')
ws_data.conditional_formatting.add(f'{current_column}1:{current_column}{row_num}', rule)
ma50d_perc_diff_column = 'I'

for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_199 = row_num - 199
        if row_num_minus_199 < 1:
            pass
        elif row_num_minus_199 >= 1:
            moving_average = f'=(SUM(B{row_num}:B{row_num_minus_199})/200)'
            ws_data[f'J{row_num}'] = moving_average
ws_data['J1'] = '200D MA'
ma200d_column = 'J'

current_column = 'K'
for row in ws_data.iter_rows(min_col=2, max_col=2, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_199 = row_num - 199
        if row_num_minus_199 < 1:
            pass
        elif row_num_minus_199 >= 1:
            percent_difference = f'=({adjclose_column}{row_num}/{ma200d_column}{row_num})-1'
            ws_data[f'{current_column}{row_num}'] = percent_difference
            ws_data[f'{current_column}{row_num}'].number_format = '0.00%'
ws_data[f'{current_column}1'] = '% diff 200D MA'
rule = ColorScaleRule(start_type='percentile', start_value=2, start_color='ff0000', mid_type='percentile', mid_value=50, mid_color='FFFFFF', end_type='percentile', end_value=98, end_color='FF00AA00')
ws_data.conditional_formatting.add(f'{current_column}1:{current_column}{row_num}', rule)
ma200d_perc_diff_column = 'K'

current_column = 'L'
for row in ws_data.iter_rows(min_col=3, max_col=3, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_19 = row_num - 19
        if row_num_minus_19 < 1:
            pass
        elif row_num_minus_19 >= 1:
            moving_volume = f'=(SUM({volume_column}{row_num}:{volume_column}{row_num_minus_19})/20)'
            ws_data[f'{current_column}{row_num}'] = moving_volume
ws_data[f'{current_column}1'] = '20D MV'
mv20d_column = 'L'

current_column = 'M'
for row in ws_data.iter_rows(min_col=3, max_col=3, values_only=False):
    for cell in row:
        cell_str = str(cell)
        coordinates = cell_str[14:-1]
        row_num = int(coordinates[1:])
        row_num_minus_19 = row_num - 19
        if row_num_minus_19 < 1:
            pass
        elif row_num_minus_19 >= 1:
            percent_difference = f'=({volume_column}{row_num}/{mv20d_column}{row_num})-1'
            ws_data[f'{current_column}{row_num}'] = percent_difference
            ws_data[f'{current_column}{row_num}'].number_format = '0.00%'
ws_data[f'{current_column}1'] = '% diff 20D MV'
rule = ColorScaleRule(start_type='percentile', start_value=2, start_color='ff0000', mid_type='percentile', mid_value=50, mid_color='FFFFFF', end_type='percentile', end_value=98, end_color='FF00AA00')
ws_data.conditional_formatting.add(f'{current_column}1:{current_column}{row_num}', rule)
mv20d_perc_diff_column = 'M'


# Volume Analysis

ws_volume_analysis = wb.create_sheet('Volume Analysis')
ws_volume_analysis['A1'] = 'Volume > 100% of 20d MV then stock continues trend for 3 days'

alpha_dict = {}
number = 0
alpha = 'abcdefghijklmnopqrstuvwxyz'
for letter in alpha.upper():
    number += 1
    alpha_dict[number] = letter

row_num = 19
for row in ws_data.iter_rows(min_col=11, max_col=11, values_only=False):
    for cell in row:
        row_num += 1
        ws_volume_analysis[f'A{row_num}'] = f"""=IF('Stock Data'!{mv20d_perc_diff_column}{row_num}>=1,TRUE,FALSE)"""
        ws_volume_analysis[f'B{row_num}'] = f"""=IF('Stock Data'!{mv20d_perc_diff_column}{row_num}>=1,'Stock Data'!{date_column}{row_num}, "")"""
        ws_volume_analysis[f'C{row_num}'] = f"""=IF('Stock Data'!{mv20d_perc_diff_column}{row_num}>=1,'Stock Data'!{adjclose_column}{row_num}, "")"""
        col_num = 3
        days_after_event = 0
        for number in range(1,11,1):
            col_num += 1
            days_after_event += 1
            ws_volume_analysis[f'{alpha_dict[col_num]}{row_num}'] = f"""=IF('Stock Data'!{mv20d_perc_diff_column}{row_num}>=1,('Stock Data'!{adjclose_column}{row_num + days_after_event}/'Stock Data'!{adjclose_column}{row_num})-1, "")"""

rule = ColorScaleRule(start_type='num', start_value=0.3, start_color='ff0000', mid_type='num', mid_value=0, mid_color='FFFFFF', end_type='num', end_value=-0.3, end_color='FF00AA00')
ws_volume_analysis.conditional_formatting.add(f'D20:M{row_num}', rule)
row_num = 0
for row in ws_volume_analysis.iter_rows(min_col=2, max_col=2):
    row_num += 1
    ws_volume_analysis[f'B{row_num}'].style = date_style
ws_volume_analysis.column_dimensions['B'].width = 11

row_num = 0
for row in ws_volume_analysis.iter_rows(min_col=4, max_col=15):
    row_num += 1
    col_num = 3
    for cell in row:
        col_num += 1
        ws_volume_analysis[f'{alpha_dict[col_num]}{row_num}'].number_format = '0.00%'

ws_volume_analysis['A3'] = f"""True"""
ws_volume_analysis['A4'] = f"""=COUNTIF(A20:A{row_num},"True")"""
ws_volume_analysis['A5'] = f"""=A4/A12"""
ws_volume_analysis['A5'].number_format = '0.00%'
ws_volume_analysis['A7'] = f"""False"""
ws_volume_analysis['A8'] = f"""=COUNTIF(A20:A{row_num},"False")"""
ws_volume_analysis['A9'] = f"""=A8/A12"""
ws_volume_analysis['A9'].number_format = '0.00%'
ws_volume_analysis['A11'] = f"""Total"""
ws_volume_analysis['A12'] = f"""=A4 + A8"""


# Drop Analysis

# use previous formula to evaluate each stock --> determine if >= 10% drop --> map out data and explain why

#200d MA tends to be elevated prior to drop --> check how many times is positive and what is average & distribution
#& how many times it is similarly elevated and does not drop
#when 200d MA is above 15% difference and 50d is above 20% check buy price

ws_drop_analysis = wb.create_sheet('Drop Analysis')
ws_drop_analysis['A1'] = 'instances where stock decreases by at least 10% in a 30d period'

row_num = 19
for row in ws_data.iter_rows(min_col=5, max_col=5, values_only=False):
    for cell in row:
        row_num += 1
        col_num = 1
        ws_drop_analysis[f'A{row_num}'] = f"""=IF('Stock Data'!{ma10d_perc_diff_column}{row_num}<=-0.07,TRUE,FALSE)"""
        col_num += 1
        ws_drop_analysis[f'B{row_num}'] = f"""=IF('Stock Data'!{ma10d_perc_diff_column}{row_num}<=-0.07,'Stock Data'!{date_column}{row_num}, "")"""
        days_after_event = 0
        days_before_event = -11
        for number in range(1,11,1):
            col_num += 1
            days_before_event += 1
            ws_drop_analysis[f'{alpha_dict[col_num]}{row_num}'] = f"""=IF('Stock Data'!{ma10d_perc_diff_column}{row_num}<=-0.07,('Stock Data'!{adjclose_column}{row_num + days_before_event}/'Stock Data'!{adjclose_column}{row_num})-1, "")"""
        col_num += 1
        ws_drop_analysis[f'{alpha_dict[col_num]}{row_num}'] = f"""=IF('Stock Data'!{ma10d_perc_diff_column}{row_num}<=-0.07,'Stock Data'!{adjclose_column}{row_num}, "")"""
        for number in range(1,11,1):
            col_num += 1
            days_after_event += 1
            ws_drop_analysis[f'{alpha_dict[col_num]}{row_num}'] = f"""=IF('Stock Data'!{ma10d_perc_diff_column}{row_num}<=-0.07,('Stock Data'!{adjclose_column}{row_num + days_after_event}/'Stock Data'!{adjclose_column}{row_num})-1, "")"""

rule = ColorScaleRule(start_type='num', start_value=0.3, start_color='ff0000', mid_type='num', mid_value=0, mid_color='FFFFFF', end_type='num', end_value=-0.3, end_color='FF00AA00')
ws_drop_analysis.conditional_formatting.add(f'C20:{alpha_dict[12]}{row_num}', rule)
ws_drop_analysis.conditional_formatting.add(f'{alpha_dict[14]}20:{alpha_dict[14+10]}{row_num}', rule)

row_num = 0
for row in ws_drop_analysis.iter_rows(min_col=2, max_col=2):
    row_num += 1
    ws_drop_analysis[f'B{row_num}'].style = date_style
ws_drop_analysis.column_dimensions['B'].width = 11

row_num = 0
for row in ws_drop_analysis.iter_rows(min_col=3, max_col=12):
    row_num += 1
    col_num = 2
    for cell in range(1,11,1):
        col_num += 1
        ws_drop_analysis[f'{alpha_dict[col_num]}{row_num}'].number_format = '0.00%'
row_num = 0
for row in ws_drop_analysis.iter_rows(min_col=14, max_col=23):
    row_num += 1
    col_num = 13
    for cell in range(1,11,1):
        col_num += 1
        ws_drop_analysis[f'{alpha_dict[col_num]}{row_num}'].number_format = '0.00%'

ws_drop_analysis['A3'] = f"""True"""
ws_drop_analysis['A4'] = f"""=COUNTIF(A20:A{row_num},"True")"""
ws_drop_analysis['A5'] = f"""=A4/A12"""
ws_drop_analysis['A5'].number_format = '0.00%'
ws_drop_analysis['A7'] = f"""False"""
ws_drop_analysis['A8'] = f"""=COUNTIF(A20:A{row_num},"False")"""
ws_drop_analysis['A9'] = f"""=A8/A12"""
ws_drop_analysis['A9'].number_format = '0.00%'
ws_drop_analysis['A11'] = f"""Total"""
ws_drop_analysis['A12'] = f"""=A4 + A8"""





ws_data.title = 'Stock Data'
dest_filename = 'tsladata.xlsx'
wb.save(filename=dest_filename)

file = r"C:\Users\Owner\PycharmProjects\03282020\tsladata.xlsx"
os.startfile(file)
